import openpyxl
wb=openpyxl.Workbook()
wb.save('test.xlsx')
#開啟檔案 命名wb 存檔檔名''


import openpyxl
wb=openpyxl.Workbook()
wb.save('tt.xlsx')
#wb.save(r'路徑')


import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
wb.create_sheet('一子甲',1)
wb.create_sheet('二子甲',0)
wb.save('test.xlsx')
#載入檔案 新增工作表 ',位置'


import openpyxl
wb=openpyxl.load_workbook('tt.xlsx')
wb.create_sheet(' 子 ',1)
wb.create_sheet(' 甲 ',0)
ws=wb['Sheet']
for cell in ws['C']:
   print(cell.value)
   wb.save('tt.xlsx')
while True:
    wb.save('tt.xlsx')
    if input()=='end':
        break


import openpyxl
wb=openpyxl.Workbook()
wb.save('test.xlsx')
for cell in ws['2']:
    print(cell.value)
    ws['a21']='=sum(a1:a20)'
    ws['b21']='=average(b1:b20)'
    ws['b22']='VALUE(b21)'
    print(ws['b22'].value) 

    
import openpyxl
wb=openpyxl.Workbook()
ws=wb['sheet1']
for cell in ws['C']:
   print(cell.value)
   wb.save('test.xlsx')


import openpyxl
wb=openpyxl.load_workbook('test.xlsx')
print(wb.sheetnames)#印全部工作表位置
wb.active=2#點選位置二
print(wb.active.title)#印出點選位置
ws=wb.active
ws.cell(row=1,column=1).value='基本電學成績'# cell:一格格 row:橫 column:直 value:數值
for i in range(1,5+1):
    #輸入5筆
    ws.cell(row=1+i,column=1).value=int(input('請輸入第'+str(i)+'筆成績:'))
ws.cell(row=7,column=1).value='=sum(a2:a6)'
ws.cell(row=1,column=2).value='資訊科技成績'
a=int(input('請輸入你的成績數量'))
for i in range(1,a+1):
    ws.cell(row=i+1,column=2).value=int(input('請輸入第'+str(i)+'筆成績:'))
ws.cell(row=a+2,column=2).value='=average(b2:b9)'
wb.save('test.xlsx')


import openpyxl
while True:
    wb=openpyxl.load_workbook('test.xlsx')
    wb.active=2
    ws=wb.active
    s=[]
    for row in ws['a2':'a6']:
        for cell in row:
            s=s+[int(cell.value)]
    print(s)
    print(sum(s))
    print(sum(s)/len(s))
    if input('請輸入0')=='0':
        break
wb.save('test.xlsx')
